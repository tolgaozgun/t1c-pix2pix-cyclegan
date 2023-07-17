"""General-purpose test script for image-to-image translation.

Once you have trained your model with train.py, you can use this script to test the model.
It will load a saved model from '--checkpoints_dir' and save the results to '--results_dir'.

It first creates model and dataset given the option. It will hard-code some parameters.
It then runs inference for '--num_test' images and save results to an HTML file.

Example (You need to train models first or download pre-trained models from our website):
    Test a CycleGAN model (both sides):
        python test.py --dataroot ./datasets/maps --name maps_cyclegan --model cycle_gan

    Test a CycleGAN model (one side only):
        python test.py --dataroot datasets/horse2zebra/testA --name horse2zebra_pretrained --model test --no_dropout

    The option '--model test' is used for generating CycleGAN results only for one side.
    This option will automatically set '--dataset_mode single', which only loads the images from one set.
    On the contrary, using '--model cycle_gan' requires loading and generating results in both directions,
    which is sometimes unnecessary. The results will be saved at ./results/.
    Use '--results_dir <directory_path_to_save_result>' to specify the results directory.

    Test a pix2pix model:
        python test.py --dataroot ./datasets/facades --name facades_pix2pix --model pix2pix --direction BtoA

See options/base_options.py and options/test_options.py for more test options.
See training and test tips at: https://github.com/junyanz/pytorch-CycleGAN-and-pix2pix/blob/master/docs/tips.md
See frequently asked questions at: https://github.com/junyanz/pytorch-CycleGAN-and-pix2pix/blob/master/docs/qa.md
"""
import os
from options.test_options import TestOptions
from data import create_dataset
from models import create_model
from util.visualizer import save_images
from util import html
import os
from torchvision.models import inception_v3
from torchvision.transforms.functional import resize, normalize
from scipy.linalg import sqrtm
import torch
import numpy as np
from numpy import cov
from numpy import trace
from numpy import iscomplexobj
from scipy.linalg import sqrtm
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    import wandb
except ImportError:
    print('Warning: wandb package cannot be found. The option "--use_wandb" will result in error.')

def calculate_fid_score(real_image, generated_image):
    """
    Calculates the Fréchet Inception Distance (FID) score between a real image and a generated image.
    
    Arguments:
    real_image -- PyTorch tensor representing the real image with shape [1, 1, 256, 256]
    generated_image -- PyTorch tensor representing the generated image with shape [1, 1, 256, 256]
    
    Returns:
    fid_score -- FID score between the real and generated images
    """
    # Convert grayscale image to RGB
    real_image_rgb = real_image.repeat(1, 3, 1, 1)
    generated_image_rgb = generated_image.repeat(1, 3, 1, 1)

    inception_model = inception_v3(pretrained=False, transform_input=False)

    # Specify the path to the .pth file containing the weights
    weights_path = "inception_v3_google-0cc3c7bd.pth"

    inception_model.load_state_dict(torch.load(weights_path))
    
    # Set the model to evaluation mode
    inception_model.eval()
    
    # Resize and normalize real and generated images
    real_image_rgb = normalize(resize(real_image_rgb, (299, 299)), (0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
    generated_image_rgb = normalize(resize(generated_image_rgb, (299, 299)), (0.5, 0.5, 0.5), (0.5, 0.5, 0.5))
    
    # Compute activations for real image
    with torch.no_grad():
        activations_real = inception_model(real_image_rgb)
        activations_real = activations_real.view(1, -1)
    
    # Compute activations for generated image
    with torch.no_grad():
        activations_generated = inception_model(generated_image_rgb)
        activations_generated = activations_generated.view(1, -1)
    
    # Calculate mean and covariance matrix for real and generated activations
    mu_real = torch.mean(activations_real, dim=0)
    mu_generated = torch.mean(activations_generated, dim=0)
    
    sigma_real = torch_cov(activations_real, rowvar=False)
    sigma_generated = torch_cov(activations_generated, rowvar=False)
    
    # Calculate FID score
    diff = mu_real - mu_generated
    fid_score = torch.trace(sigma_real + sigma_generated - 2 * sqrtm(sigma_real @ sigma_generated))
    fid_score += torch.matmul(diff, diff)
    fid_score = fid_score.real.item()
    
    return fid_score

def torch_cov(m, rowvar=False):
    """
    Estimate a covariance matrix given data.
    
    Arguments:
    m -- PyTorch tensor representing the data
    rowvar -- If True, then each row represents a variable (default: False)
    
    Returns:
    cov -- Covariance matrix of the input data
    """
    if m.dim() > 2:
        raise ValueError('m has more than 2 dimensions')
    if m.dim() < 2:
        m = m.view(1, -1)
    if not rowvar and m.size(0) != 1:
        m = m.t()
    fact = 1.0 / (m.size(1) - 1)
    m -= torch.mean(m, dim=1, keepdim=True)
    mt = m.t()
    return fact * m @ mt



def save_fid_to_file(epoch, fid_score):
    # Define the file path
    file_path = "fid_scores.xlsx"

    # Create a new workbook or load an existing one
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active worksheet or create a new one
    worksheet = workbook.active if workbook.sheetnames else workbook.create_sheet()

    # Find the next empty row
    next_row = worksheet.max_row + 1

    # Write the data to the worksheet
    worksheet.cell(row=next_row, column=1, value=epoch)
    worksheet.cell(row=next_row, column=2, value=fid_score)

    # Save the workbook to the file
    workbook.save(file_path)


# calculate frechet inception distance
def calculate_fid(act1, act2):
    # calculate mean and covariance statistics
    mu1, sigma1 = act1.mean(axis=0), cov(act1, rowvar=False)
    mu2, sigma2 = act2.mean(axis=0), cov(act2, rowvar=False)
    # calculate sum squared difference between means
    ssdiff = np.sum((mu1 - mu2)**2.0)
    # calculate sqrt of product between cov
    covmean = sqrtm(sigma1.dot(sigma2))
    # check and correct imaginary numbers from sqrt
    if iscomplexobj(covmean):
        covmean = covmean.real
    # calculate score
    fid = ssdiff + trace(sigma1 + sigma2 - 2.0 * covmean)
    return fid

if __name__ == '__main__':
    opt = TestOptions().parse()  # get test options
    # hard-code some parameters for test
    opt.num_threads = 0   # test code only supports num_threads = 0
    opt.batch_size = 1    # test code only supports batch_size = 1
    opt.serial_batches = True  # disable data shuffling; comment this line if results on randomly chosen images are needed.
    opt.no_flip = True    # no flip; comment this line if results on flipped images are needed.
    opt.display_id = -1   # no visdom display; the test code saves the results to a HTML file.
    dataset = create_dataset(opt)  # create a dataset given opt.dataset_mode and other options
    model = create_model(opt)      # create a model given opt.model and other options
    model.setup(opt)               # regular setup: load and print networks; create schedulers

    # initialize logger
    if opt.use_wandb:
        wandb_run = wandb.init(project=opt.wandb_project_name, name=opt.name, config=opt) if not wandb.run else wandb.run
        wandb_run._label(repo='CycleGAN-and-pix2pix')

    # create a website
    web_dir = os.path.join(opt.results_dir, opt.name, '{}_{}'.format(opt.phase, opt.epoch))  # define the website directory
    if opt.load_iter > 0:  # load_iter is 0 by default
        web_dir = '{:s}_iter{:d}'.format(web_dir, opt.load_iter)
    print('creating web directory', web_dir)
    webpage = html.HTML(web_dir, 'Experiment = %s, Phase = %s, Epoch = %s' % (opt.name, opt.phase, opt.epoch))
    # test with eval mode. This only affects layers like batchnorm and dropout.
    # For [pix2pix]: we use batchnorm and dropout in the original pix2pix. You can experiment it with and without eval() mode.
    # For [CycleGAN]: It should not affect CycleGAN as CycleGAN uses instancenorm without dropout.

    fid_scores = []
    if opt.eval:
        model.eval()
    for i, data in enumerate(dataset):
        if i >= opt.num_test:  # only apply our model to opt.num_test images.
            break
        model.set_input(data)  # unpack data from data loader
        model.test()           # run inference
        visuals = model.get_current_visuals()  # get image results
        fid = calculate_fid_score(visuals['fake_B'], visuals['real_B'])
        fid_scores.append(fid)

        img_path = model.get_image_paths()     # get image paths
        if i % 5 == 0:  # save images to an HTML file
            print('processing (%04d)-th image... %s' % (i, img_path))
        save_images(webpage, visuals, fid, img_path, aspect_ratio=opt.aspect_ratio, width=opt.display_winsize, use_wandb=opt.use_wandb)
    mean_fid = np.mean(fid)
    print(f"Fid_scores array length: {len(fid_scores)}")
    print(f"Mean fid: {mean_fid}")
    save_fid_to_file(opt.epoch, mean_fid)
    webpage.save()  # save the HTML


